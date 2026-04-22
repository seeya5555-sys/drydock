"""
Fleet Dry Dock Manager — Flask Backend
=======================================
Run:
    pip install flask
    python app.py

Open: http://localhost:5000
"""

import sqlite3, os, io, json, hashlib, secrets, gzip
from flask import Flask, g, jsonify, request, render_template, abort, send_file, session, redirect, url_for

app = Flask(__name__)
app.config["DATABASE"] = os.path.join(app.instance_path, "fleet.db")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024   # 50 MB
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "drydock-secret-key-2026")
app.config["PERMANENT_SESSION_LIFETIME"] = 60 * 60 * 24 * 7  # 7일
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0  # 캐시버스팅 있으니 0
os.makedirs(app.instance_path, exist_ok=True)

# ── MCP / OAuth 디버그 로거 ───────────────────────────────────
@app.before_request
def _mcp_debug_log():
    p = request.path
    if any(x in p for x in ('/mcp', '/oauth', '/.well-known')):
        auth = 'Y' if request.headers.get('Authorization') else 'N'
        ua   = request.headers.get('User-Agent','')[:80]
        ct   = request.headers.get('Content-Type','')
        print(f"[MCP-IN ] {request.method:6s} {p:40s} auth={auth} ct={ct} ua={ua}", flush=True)

@app.after_request
def _mcp_debug_log_out(resp):
    p = request.path
    if any(x in p for x in ('/mcp', '/oauth', '/.well-known')):
        print(f"[MCP-OUT] {request.method:6s} {p:40s} -> {resp.status_code}", flush=True)
    return resp

# ── Gzip 압축 ─────────────────────────────────────────────────
@app.after_request
def compress_response(response):
    if (response.status_code < 200 or response.status_code >= 300
            or response.direct_passthrough):
        return response
    accept = request.headers.get('Accept-Encoding', '')
    if 'gzip' not in accept:
        return response
    content_type = response.content_type
    if not any(t in content_type for t in ('json', 'javascript', 'css', 'html', 'text')):
        return response
    data = response.get_data()
    if len(data) < 500:
        return response
    compressed = gzip.compress(data, compresslevel=6)
    if len(compressed) >= len(data):
        return response
    response.set_data(compressed)
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(compressed)
    return response


# ── DB helpers ────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"],
                               detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db: db.close()

def init_db():
    db = get_db()
    with open(os.path.join(os.path.dirname(__file__), "schema.sql"), encoding='utf-8') as f:
        db.executescript(f.read())
    db.commit()

def rows(sql, *args):
    return [dict(r) for r in get_db().execute(sql, args).fetchall()]

def row(sql, *args):
    r = get_db().execute(sql, args).fetchone()
    return dict(r) if r else None

with app.app_context():
    init_db()
    db = get_db()
    # STORE 섹션 수동 Budget/Consumed 테이블
    db.execute("""CREATE TABLE IF NOT EXISTS vessel_sec_budget (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vessel_id TEXT NOT NULL,
        category TEXT NOT NULL,
        section TEXT NOT NULL,
        budget REAL DEFAULT 0,
        consumed REAL DEFAULT 0,
        UNIQUE(vessel_id, category, section)
    )""")
    # Shipyard D/C 요율 컬럼 추가 (없으면)
    try:
        db.execute("ALTER TABLE vessels ADD COLUMN dc_rate REAL DEFAULT 0")
    except:
        pass
    # 신규 날짜 컬럼 추가 (없으면)
    for col in ['berthing_date', 'departure_date']:
        try:
            db.execute(f"ALTER TABLE vessels ADD COLUMN {col} TEXT")
        except:
            pass
    # steel_repair REV2 컬럼 추가 (기존 데이터 보존)
    for col, default in [
        ('position_tank',   "''"), ('frame_no',       "''"),
        ('location_detail', "''"), ('type',            "''"),
        ('length_l',        "''"), ('width_w',         "''"),
        ('thickness_t',     "''"), ('steel_grade',     "''"),
        ('new_weight',      "''"), ('space_type',      "'Open Space'"),
        ('shape',           "'Flat'"), ('test_required', "'None'"),
        ('staging_m3',      "''"), ('est_cost',        "''"),
        ('actual_charged',  "''"),
    ]:
        try:
            db.execute(f"ALTER TABLE steel_repair ADD COLUMN {col} TEXT DEFAULT {default}")
        except:
            pass
    # pipe_repair 테이블 생성
    db.execute("""CREATE TABLE IF NOT EXISTS pipe_repair (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        vessel_id       TEXT NOT NULL REFERENCES vessels(id) ON DELETE CASCADE,
        no              TEXT,
        system_line     TEXT,
        position_tank   TEXT,
        frame_no        TEXT,
        location_detail TEXT,
        pipe_od         TEXT,
        schedule        TEXT DEFAULT 'Sch40',
        material        TEXT DEFAULT 'Carbon Steel',
        length_m        TEXT,
        bend_qty        TEXT,
        flange_qty      TEXT,
        valve_type      TEXT DEFAULT 'None',
        valve_size      TEXT,
        valve_qty       TEXT,
        space_type      TEXT DEFAULT 'Open Deck / Open Space',
        removal_refit   TEXT DEFAULT 'Not Required',
        priority        TEXT DEFAULT 'Normal',
        status          TEXT DEFAULT 'Not Started',
        start_date      TEXT,
        completion_date TEXT,
        est_cost        TEXT,
        actual_charged  TEXT,
        remark          TEXT,
        last_updated    TEXT DEFAULT (datetime('now'))
    )""")
    # pipe_repair description 컬럼 추가
    try:
        db.execute("ALTER TABLE pipe_repair ADD COLUMN description TEXT DEFAULT ''")
    except:
        pass
    # Tank layout per vessel
    db.execute("""CREATE TABLE IF NOT EXISTS vessel_tank_layout (
        vessel_id   TEXT PRIMARY KEY REFERENCES vessels(id) ON DELETE CASCADE,
        layout_json TEXT NOT NULL,
        updated_at  TEXT DEFAULT (datetime('now'))
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS vessel_wps_criteria (
        vessel_id      TEXT PRIMARY KEY REFERENCES vessels(id) ON DELETE CASCADE,
        criteria_json  TEXT NOT NULL,
        updated_at     TEXT DEFAULT (datetime('now'))
    )""")
    db.commit()
    db = get_db()
    # Documents 테이블 생성
    db.execute("""CREATE TABLE IF NOT EXISTS vessel_documents (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        vessel_id   TEXT NOT NULL REFERENCES vessels(id) ON DELETE CASCADE,
        doc_type    TEXT NOT NULL,
        filename    TEXT NOT NULL,
        filesize    INTEGER,
        mimetype    TEXT,
        data        BLOB NOT NULL,
        uploaded_at TEXT DEFAULT (datetime('now'))
    )""")
    db.commit()
    db = get_db()
    # 사용자 테이블 생성 (role, vessels 포함)
    db.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'editor',
        vessels TEXT DEFAULT '[]',
        created_at DATETIME DEFAULT (datetime('now'))
    )""")
    # 기존 테이블에 컬럼 추가 (마이그레이션)
    for col, default in [('role',"'editor'"), ('vessels',"'[]'")]:
        try:
            db.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT DEFAULT {default}")
        except: pass
    # 기본 admin 계정
    def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()
    if not db.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
        db.execute("INSERT INTO users(username,password_hash,role,vessels) VALUES(?,?,?,?)",
                   ('admin', hash_pw('drydock2026'), 'admin', '[]'))
    else:
        db.execute("UPDATE users SET role='admin' WHERE username='admin'")
    db.commit()

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def get_current_user():
    if not session.get('logged_in'): return None
    return get_db().execute("SELECT * FROM users WHERE username=?",
                            (session['username'],)).fetchone()

def is_admin():
    u = get_current_user()
    return u and u['role'] == 'admin'

def can_access_vessel(vid):
    u = get_current_user()
    if not u: return False
    if u['role'] == 'admin': return True
    allowed = json.loads(u['vessels'] or '[]')
    return vid in allowed

def is_viewer():
    u = get_current_user()
    return u and u['role'] == 'viewer'

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def viewer_forbidden(f):
    """읽기 전용(viewer) 계정은 쓰기 불가"""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if is_viewer():
            return jsonify({"error": "읽기 전용 계정입니다"}), 403
        return f(*args, **kwargs)
    return decorated


# ── JSON column helpers ───────────────────────────────────────
def je(val):
    """list → JSON string"""
    return json.dumps(val, ensure_ascii=False) if isinstance(val, list) else (val or '[]')

def jd(val):
    """JSON string → list"""
    if not val: return []
    try: return json.loads(val)
    except: return []


# ── Row normalizers (DB snake_case → frontend camelCase) ──────
def to_vessel(r):
    return {"id": r["id"], "name": r["name"] or "",
            "type": r["type"] or "", "imo": r["imo"] or "",
            "shipyard": r["shipyard"] or "",
            "classSociety": r["class_society"] or "",
            "berthingDate": r["berthing_date"] or "",
            "dockIn": r["dock_in"] or "", "dockOut": r["dock_out"] or "",
            "departureDate": r["departure_date"] or "",
            "duration": r["duration"] or "", "grt": r["grt"] or "",
            "dcRate": r["dc_rate"] if r["dc_rate"] is not None else 0}

def to_job(r):
    return {"_id": r["id"], "number": r["number"] or "",
            "section": r["section"] or "GENERAL",
            "category": r["category"] or "Shipyard",
            "description": r["description"] or "",
            "vendor": r["vendor"] or "",
            "budget": r["budget"] or 0, "consumption": r["consumption"] or 0,
            "start_date": r["start_date"] or "", "end_date": r["end_date"] or "",
            "completion": r["completion"] or 0,
            "remarks": jd(r["remarks"])}

def to_class(r):
    return {"_id": r["id"], "no": r["no"] or "",
            "finding": r["finding"] or "", "description": r["description"] or "",
            "actions": jd(r["actions"]),
            "by": r["responsible"] or "Crew",
            "open_date": r["open_date"] or "", "close_date": r["close_date"] or "",
            "status": r["status"] or "Open", "priority": r["priority"] or "Normal"}

def to_disc(r):
    return {"_id": r["id"], "no": r["no"] or "",
            "date": r["date"] or "", "time_of_day": r["time_of_day"] or "",
            "item": r["item"] or "", "description": r["description"] or "",
            "actions": jd(r["actions"]),
            "status": r["status"] or "Open", "priority": r["priority"] or "Normal"}


# ── Frontend ──────────────────────────────────────────────────
@app.route("/login", methods=["GET"])
def login_page():
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def do_login():
    data = request.get_json(force=True)
    username = data.get("username","").strip()
    password = data.get("password","")
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username=? AND password_hash=?",
                      (username, hash_pw(password))).fetchone()
    if user:
        session.permanent = True
        session['logged_in'] = True
        session['username'] = username
        session['role'] = user['role'] or 'editor'
        session['vessels'] = user['vessels'] or '[]'
        return jsonify({"ok": True, "role": session['role']})
    return jsonify({"ok": False, "error": "아이디 또는 비밀번호가 틀렸습니다"}), 401

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route("/api/auth/me", methods=["GET"])
@login_required
def get_me():
    u = get_current_user()
    if not u: return jsonify({"error": "Not found"}), 404
    return jsonify({"username": u['username'], "role": u['role'],
                    "vessels": json.loads(u['vessels'] or '[]')})

@app.route("/api/auth/users", methods=["GET"])
@login_required
def get_users():
    users = rows("SELECT id, username, role, vessels, created_at FROM users ORDER BY id")
    return jsonify(users)

@app.route("/api/auth/users", methods=["POST"])
@login_required
def add_user():
    if not is_admin():
        return jsonify({"error": "관리자만 계정을 생성할 수 있습니다"}), 403
    data = request.get_json(force=True)
    username = data.get("username","").strip()
    password = data.get("password","")
    role = data.get("role","editor")
    vessels = json.dumps(data.get("vessels", []))
    if not username or not password:
        return jsonify({"error": "아이디와 비밀번호를 입력하세요"}), 400
    if role not in ('admin','editor','viewer'):
        return jsonify({"error": "유효하지 않은 역할입니다"}), 400
    try:
        get_db().execute("INSERT INTO users(username,password_hash,role,vessels) VALUES(?,?,?,?)",
                         (username, hash_pw(password), role, vessels))
        get_db().commit()
        return jsonify({"ok": True})
    except:
        return jsonify({"error": "이미 존재하는 아이디입니다"}), 409

@app.route("/api/auth/users/<int:uid>", methods=["DELETE"])
@login_required
def delete_user(uid):
    if not is_admin():
        return jsonify({"error": "관리자만 계정을 삭제할 수 있습니다"}), 403
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user: return jsonify({"error": "사용자 없음"}), 404
    if user["username"] == "admin":
        return jsonify({"error": "admin 계정은 삭제할 수 없습니다"}), 403
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/auth/users/<int:uid>", methods=["PUT"])
@login_required
def update_user(uid):
    if not is_admin():
        return jsonify({"error": "관리자만 계정을 수정할 수 있습니다"}), 403
    data = request.get_json(force=True)
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user: return jsonify({"error": "사용자 없음"}), 404
    vessels = json.dumps(data.get("vessels", json.loads(user['vessels'] or '[]')))
    role = data.get("role", user['role'])
    db.execute("UPDATE users SET role=?, vessels=? WHERE id=?", (role, vessels, uid))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/auth/password", methods=["PUT"])
@login_required
def change_password():
    data = request.get_json(force=True)
    username = session.get('username')
    old_pw = data.get("old_password","")
    new_pw = data.get("new_password","")
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username=? AND password_hash=?",
                      (username, hash_pw(old_pw))).fetchone()
    if not user:
        return jsonify({"error": "현재 비밀번호가 틀렸습니다"}), 401
    db.execute("UPDATE users SET password_hash=? WHERE username=?",
               (hash_pw(new_pw), username))
    db.commit()
    return jsonify({"ok": True})

@app.route("/")
@login_required
def index():
    import time
    return render_template("index.html", version=int(time.time()))


# ══════════════════════════════════════════════════════════════
# VESSELS
# ══════════════════════════════════════════════════════════════

@app.route("/api/vessels", methods=["GET"])
def get_vessels():
    return jsonify([to_vessel(r) for r in
                    get_db().execute("SELECT * FROM vessels ORDER BY created_at").fetchall()])

@app.route("/api/vessels", methods=["POST"])
def create_vessel():
    d = request.get_json(force=True)
    vid = d.get("id") or f"v_{os.urandom(4).hex()}"
    db = get_db()
    db.execute(
        "INSERT INTO vessels(id,name,type,imo,shipyard,class_society,berthing_date,dock_in,dock_out,departure_date,duration,grt)"
        " VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        (vid, d.get("name",""), d.get("type",""), d.get("imo",""),
         d.get("shipyard",""), d.get("classSociety",""),
         d.get("berthingDate") or None,
         d.get("dockIn") or None, d.get("dockOut") or None,
         d.get("departureDate") or None,
         d.get("duration") or None, d.get("grt","")))
    db.commit()
    return jsonify(to_vessel(row("SELECT * FROM vessels WHERE id=?", vid))), 201

@app.route("/api/vessels/<vid>", methods=["PUT"])
def update_vessel(vid):
    d = request.get_json(force=True)
    db = get_db()
    db.execute(
        "UPDATE vessels SET name=?,type=?,imo=?,shipyard=?,class_society=?,"
        "berthing_date=?,dock_in=?,dock_out=?,departure_date=?,duration=?,grt=? WHERE id=?",
        (d.get("name",""), d.get("type",""), d.get("imo",""),
         d.get("shipyard",""), d.get("classSociety",""),
         d.get("berthingDate") or None,
         d.get("dockIn") or None, d.get("dockOut") or None,
         d.get("departureDate") or None,
         d.get("duration") or None, d.get("grt",""), vid))
    db.commit()
    v = row("SELECT * FROM vessels WHERE id=?", vid)
    return jsonify(to_vessel(v)) if v else abort(404)

@app.route("/api/vessels/<vid>", methods=["DELETE"])
def delete_vessel(vid):
    db = get_db()
    db.execute("DELETE FROM vessels WHERE id=?", (vid,))
    db.commit()
    return jsonify({"deleted": vid})

@app.route("/api/vessels/<vid>/dc_rate", methods=["PUT"])
@login_required
@viewer_forbidden
def set_dc_rate(vid):
    d = request.get_json(force=True)
    rate = float(d.get("dcRate", 0) or 0)
    db = get_db()
    db.execute("UPDATE vessels SET dc_rate=? WHERE id=?", (rate, vid))
    db.commit()
    return jsonify({"ok": True, "dcRate": rate})

@app.route("/api/vessels/<vid>/sec_budget", methods=["GET"])
@login_required
def get_sec_budget(vid):
    data = rows("SELECT category, section, budget, consumed FROM vessel_sec_budget WHERE vessel_id=?", vid)
    return jsonify(data)

@app.route("/api/vessels/<vid>/sec_budget", methods=["PUT"])
@login_required
@viewer_forbidden
def set_sec_budget(vid):
    d = request.get_json(force=True)
    cat = d.get("category","")
    sec = d.get("section","")
    budget   = float(d.get("budget",0) or 0)
    consumed = float(d.get("consumed",0) or 0)
    db = get_db()
    db.execute("""INSERT INTO vessel_sec_budget(vessel_id,category,section,budget,consumed)
                  VALUES(?,?,?,?,?)
                  ON CONFLICT(vessel_id,category,section) DO UPDATE SET budget=?,consumed=?""",
               (vid, cat, sec, budget, consumed, budget, consumed))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/fleet/summary")
@login_required
def fleet_summary():
    db = get_db()
    u = get_current_user()
    allowed = None if (u and u['role']=='admin') else json.loads(u['vessels'] or '[]') if u else []
    # 전체 데이터 한 번에 로드 (vessel별 반복 쿼리 제거)
    all_jobs        = {}
    all_class       = {}
    all_disc        = {}
    all_attach      = {}
    all_secbudget   = {}

    for r in db.execute("SELECT * FROM jobs ORDER BY vessel_id, id").fetchall():
        all_jobs.setdefault(r["vessel_id"], []).append(to_job(r))
    for r in db.execute("SELECT * FROM class_items ORDER BY vessel_id, id").fetchall():
        all_class.setdefault(r["vessel_id"], []).append(to_class(r))
    for r in db.execute("SELECT * FROM discussions ORDER BY vessel_id, date, id").fetchall():
        all_disc.setdefault(r["vessel_id"], []).append(to_disc(r))
    for r in db.execute("SELECT vessel_id, ref_type, ref_id FROM attachments GROUP BY vessel_id, ref_type, ref_id").fetchall():
        all_attach.setdefault(r["vessel_id"], []).append({"ref_type": r["ref_type"], "ref_id": r["ref_id"]})
    for r in db.execute("SELECT vessel_id, category, section, budget, consumed FROM vessel_sec_budget").fetchall():
        all_secbudget.setdefault(r["vessel_id"], []).append({"category": r["category"], "section": r["section"], "budget": r["budget"], "consumed": r["consumed"]})

    result = []
    for v in db.execute("SELECT * FROM vessels ORDER BY created_at").fetchall():
        vid = v["id"]
        if allowed is not None and vid not in allowed:
            continue
        result.append({
            "info":        to_vessel(v),
            "jobs":        all_jobs.get(vid, []),
            "classItems":  all_class.get(vid, []),
            "discussions": all_disc.get(vid, []),
            "attachments": all_attach.get(vid, []),
            "secBudget":   all_secbudget.get(vid, []),
        })
    return jsonify(result)


# ══════════════════════════════════════════════════════════════
# JOBS
# ══════════════════════════════════════════════════════════════

@app.route("/api/vessels/<vid>/jobs/csv", methods=["POST"])
def upload_jobs_csv(vid):
    """CSV 파일로 Job 일괄 등록
    컬럼: number, section, category, description, vendor, budget, start_date, end_date, completion
    """
    import csv, io as _io

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    if not f.filename.endswith(".csv"):
        return jsonify({"error": "CSV 파일만 업로드 가능합니다"}), 400

    stream = _io.StringIO(f.stream.read().decode("utf-8-sig"))
    reader = csv.DictReader(stream)

    required = {"number", "section", "category", "description"}
    if not required.issubset({c.strip().lower() for c in (reader.fieldnames or [])}):
        return jsonify({"error": f"필수 컬럼 누락: {required}"}), 400

    db = get_db()
    inserted = 0
    updated = 0
    errors = []

    # 기존 번호 목록 미리 로드
    existing = {r["number"]: r["id"] for r in db.execute(
        "SELECT id, number FROM jobs WHERE vessel_id=?", (vid,)).fetchall()}

    for i, row_data in enumerate(reader, start=2):
        r = {k.strip().lower(): v.strip() for k, v in row_data.items()}
        raw_budget = r.get("budget", "0") or "0"
        clean_budget = raw_budget.replace("$", "").replace(",", "").strip()

        def clean_date(v):
            v = (v or "").strip()
            return v if v else None

        raw_comp = r.get("completion", "0") or "0"
        try:
            completion = int(float(raw_comp.replace("%","").strip()))
        except:
            completion = 0

        number = r.get("number", "").strip()
        if not number:  # 빈 행 무시
            continue
        start_date = clean_date(r.get("start_date"))
        end_date   = clean_date(r.get("end_date"))
        budget     = float(clean_budget or 0)

        try:
            if number in existing:
                # 기존 번호 → 날짜/Budget/Consumed만 업데이트 (입력값이 있을 때만)
                updates = []
                params  = []
                if start_date:
                    updates.append("start_date=?"); params.append(start_date)
                if end_date:
                    updates.append("end_date=?"); params.append(end_date)
                if budget > 0:
                    updates.append("budget=?"); params.append(budget)
                if completion > 0:
                    updates.append("completion=?"); params.append(completion)
                if updates:
                    params.append(existing[number])
                    db.execute(f"UPDATE jobs SET {','.join(updates)} WHERE id=?", params)
                    updated += 1
            else:
                # 신규 번호 → INSERT
                db.execute(
                    "INSERT INTO jobs(vessel_id, number, section, category, description, "
                    "vendor, budget, consumption, start_date, end_date, completion, remarks) "
                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        vid,
                        number,
                        r.get("section", "GENERAL").upper(),
                        r.get("category", "Shipyard"),
                        r.get("description", ""),
                        r.get("vendor", ""),
                        budget,
                        0,
                        start_date,
                        end_date,
                        completion,
                        "[]",
                    )
                )
                inserted += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    db.commit()
    return jsonify({"inserted": inserted, "updated": updated, "errors": errors}), 201

@app.route("/api/vessels/<vid>/jobs", methods=["GET"])
def get_jobs(vid):
    return jsonify([to_job(r) for r in
                    get_db().execute("SELECT * FROM jobs WHERE vessel_id=? ORDER BY id", (vid,)).fetchall()])

@app.route("/api/vessels/<vid>/jobs", methods=["POST"])
def create_job(vid):
    d = request.get_json(force=True)
    db = get_db()
    cur = db.execute(
        "INSERT INTO jobs(vessel_id,number,section,category,description,vendor,"
        "budget,consumption,start_date,end_date,completion,remarks) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        (vid, d.get("number",""), d.get("section","GENERAL"), d.get("category","Shipyard"),
         d.get("description",""), d.get("vendor",""),
         d.get("budget",0), d.get("consumption",0),
         d.get("start_date") or None, d.get("end_date") or None,
         d.get("completion",0), je(d.get("remarks",[]))))
    db.commit()
    return jsonify(to_job(row("SELECT * FROM jobs WHERE id=?", cur.lastrowid))), 201

@app.route("/api/jobs/<int:jid>", methods=["PUT"])
def update_job(jid):
    d = request.get_json(force=True)
    db = get_db()
    db.execute(
        "UPDATE jobs SET number=?,section=?,category=?,description=?,vendor=?,"
        "budget=?,consumption=?,start_date=?,end_date=?,completion=?,remarks=?,"
        "updated_at=datetime('now') WHERE id=?",
        (d.get("number",""), d.get("section","GENERAL"), d.get("category","Shipyard"),
         d.get("description",""), d.get("vendor",""),
         d.get("budget",0), d.get("consumption",0),
         d.get("start_date") or None, d.get("end_date") or None,
         d.get("completion",0), je(d.get("remarks",[])), jid))
    db.commit()
    return jsonify(to_job(row("SELECT * FROM jobs WHERE id=?", jid)))

@app.route("/api/jobs/<int:jid>", methods=["DELETE"])
@login_required
@viewer_forbidden
def delete_job(jid):
    db = get_db()
    db.execute("DELETE FROM jobs WHERE id=?", (jid,))
    db.commit()
    return jsonify({"deleted": jid})

@app.route("/api/vessels/<vid>/jobs/bulk", methods=["PUT"])
def bulk_jobs(vid):
    db = get_db()
    db.execute("DELETE FROM jobs WHERE vessel_id=?", (vid,))
    for j in request.get_json(force=True):
        db.execute(
            "INSERT INTO jobs(vessel_id,number,section,category,description,vendor,"
            "budget,consumption,start_date,end_date,completion,remarks) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (vid, j.get("number",""), j.get("section","GENERAL"), j.get("category","Shipyard"),
             j.get("description",""), j.get("vendor",""),
             j.get("budget",0), j.get("consumption",0),
             j.get("start_date") or None, j.get("end_date") or None,
             j.get("completion",0), je(j.get("remarks",[]))))
    db.commit()
    return jsonify([to_job(r) for r in
                    get_db().execute("SELECT * FROM jobs WHERE vessel_id=? ORDER BY id", (vid,)).fetchall()])


# ══════════════════════════════════════════════════════════════
# CLASS ITEMS
# ══════════════════════════════════════════════════════════════

@app.route("/api/vessels/<vid>/class_items", methods=["GET"])
def get_class_items(vid):
    return jsonify([to_class(r) for r in
                    get_db().execute("SELECT * FROM class_items WHERE vessel_id=? ORDER BY id", (vid,)).fetchall()])

@app.route("/api/vessels/<vid>/class_items", methods=["POST"])
def create_class_item(vid):
    d = request.get_json(force=True)
    db = get_db()
    cur = db.execute(
        "INSERT INTO class_items(vessel_id,no,finding,description,actions,responsible,"
        "open_date,close_date,status,priority) VALUES(?,?,?,?,?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("finding",""), d.get("description",""),
         je(d.get("actions",[])), d.get("by","Crew"),
         d.get("open_date") or None, d.get("close_date") or None,
         d.get("status","Open"), d.get("priority","Normal")))
    db.commit()
    return jsonify(to_class(row("SELECT * FROM class_items WHERE id=?", cur.lastrowid))), 201

@app.route("/api/class_items/<int:cid>", methods=["PUT"])
def update_class_item(cid):
    d = request.get_json(force=True)
    db = get_db()
    db.execute(
        "UPDATE class_items SET no=?,finding=?,description=?,actions=?,responsible=?,"
        "open_date=?,close_date=?,status=?,priority=?,updated_at=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("finding",""), d.get("description",""),
         je(d.get("actions",[])), d.get("by","Crew"),
         d.get("open_date") or None, d.get("close_date") or None,
         d.get("status","Open"), d.get("priority","Normal"), cid))
    db.commit()
    return jsonify(to_class(row("SELECT * FROM class_items WHERE id=?", cid)))

@app.route("/api/class_items/<int:cid>", methods=["DELETE"])
def delete_class_item(cid):
    db = get_db()
    db.execute("DELETE FROM class_items WHERE id=?", (cid,))
    db.commit()
    return jsonify({"deleted": cid})

@app.route("/api/vessels/<vid>/class_items/bulk", methods=["PUT"])
def bulk_class_items(vid):
    db = get_db()
    db.execute("DELETE FROM class_items WHERE vessel_id=?", (vid,))
    for item in request.get_json(force=True):
        db.execute(
            "INSERT INTO class_items(vessel_id,no,finding,description,actions,responsible,"
            "open_date,close_date,status,priority) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (vid, item.get("no",""), item.get("finding",""), item.get("description",""),
             je(item.get("actions",[])), item.get("by","Crew"),
             item.get("open_date") or None, item.get("close_date") or None,
             item.get("status","Open"), item.get("priority","Normal")))
    db.commit()
    return jsonify([to_class(r) for r in
                    get_db().execute("SELECT * FROM class_items WHERE vessel_id=? ORDER BY id", (vid,)).fetchall()])


# ══════════════════════════════════════════════════════════════
# DISCUSSIONS
# ══════════════════════════════════════════════════════════════

@app.route("/api/vessels/<vid>/discussions", methods=["GET"])
def get_discussions(vid):
    return jsonify([to_disc(r) for r in
                    get_db().execute("SELECT * FROM discussions WHERE vessel_id=? ORDER BY date,id", (vid,)).fetchall()])

@app.route("/api/vessels/<vid>/discussions", methods=["POST"])
def create_discussion(vid):
    d = request.get_json(force=True)
    db = get_db()
    cur = db.execute(
        "INSERT INTO discussions(vessel_id,no,date,time_of_day,item,description,actions,status,priority)"
        " VALUES(?,?,?,?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("date") or None, d.get("time_of_day",""),
         d.get("item",""), d.get("description",""),
         je(d.get("actions",[])), d.get("status","Open"), d.get("priority","Normal")))
    db.commit()
    return jsonify(to_disc(row("SELECT * FROM discussions WHERE id=?", cur.lastrowid))), 201

@app.route("/api/discussions/<int:did>", methods=["PUT"])
def update_discussion(did):
    d = request.get_json(force=True)
    db = get_db()
    db.execute(
        "UPDATE discussions SET no=?,date=?,time_of_day=?,item=?,description=?,actions=?,"
        "status=?,priority=?,updated_at=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("date") or None, d.get("time_of_day",""),
         d.get("item",""), d.get("description",""),
         je(d.get("actions",[])), d.get("status","Open"), d.get("priority","Normal"), did))
    db.commit()
    return jsonify(to_disc(row("SELECT * FROM discussions WHERE id=?", did)))

@app.route("/api/discussions/<int:did>", methods=["DELETE"])
def delete_discussion(did):
    db = get_db()
    db.execute("DELETE FROM discussions WHERE id=?", (did,))
    db.commit()
    return jsonify({"deleted": did})

@app.route("/api/vessels/<vid>/discussions/bulk", methods=["PUT"])
def bulk_discussions(vid):
    db = get_db()
    items = request.get_json(force=True)

    # 전송된 _id 목록
    incoming_ids = [int(item['_id']) for item in items if item.get('_id')]

    # DB에 있지만 전송 목록에 없는 항목 삭제 (삭제된 row)
    if incoming_ids:
        placeholders = ','.join('?' * len(incoming_ids))
        db.execute(f"DELETE FROM discussions WHERE vessel_id=? AND id NOT IN ({placeholders})",
                   [vid] + incoming_ids)
    else:
        db.execute("DELETE FROM discussions WHERE vessel_id=?", (vid,))

    result_ids = []
    for item in items:
        did = item.get('_id')
        vals = (item.get("no",""), item.get("date") or None, item.get("time_of_day",""),
                item.get("item",""), item.get("description",""),
                je(item.get("actions",[])), item.get("status","Open"), item.get("priority","Normal"))
        if did:
            # 기존 row → UPDATE (ID 보존 → 첨부파일 ref_id 유지)
            db.execute(
                "UPDATE discussions SET no=?,date=?,time_of_day=?,item=?,description=?,"
                "actions=?,status=?,priority=?,updated_at=datetime('now') WHERE id=? AND vessel_id=?",
                vals + (int(did), vid))
            result_ids.append(int(did))
        else:
            # 신규 row → INSERT
            cur = db.execute(
                "INSERT INTO discussions(vessel_id,no,date,time_of_day,item,description,actions,status,priority)"
                " VALUES(?,?,?,?,?,?,?,?,?)",
                (vid,) + vals)
            result_ids.append(cur.lastrowid)

    db.commit()
    return jsonify([to_disc(r) for r in
                    get_db().execute("SELECT * FROM discussions WHERE vessel_id=? ORDER BY date,id", (vid,)).fetchall()])


# ══════════════════════════════════════════════════════════════
# FILE ATTACHMENTS
# ══════════════════════════════════════════════════════════════

@app.route("/api/vessels/<vid>/attachments/<ref_type>/<int:ref_id>", methods=["GET"])
def get_attachments(vid, ref_type, ref_id):
    return jsonify(rows(
        "SELECT id,filename,filesize,mimetype,uploaded_at FROM attachments"
        " WHERE vessel_id=? AND ref_type=? AND ref_id=? ORDER BY id",
        vid, ref_type, ref_id))

@app.route("/api/vessels/<vid>/attachments/<ref_type>/<int:ref_id>", methods=["POST"])
def upload_attachment(vid, ref_type, ref_id):
    if "files" not in request.files:
        return jsonify({"error": "No files"}), 400
    db = get_db()
    uploaded = []
    for f in request.files.getlist("files"):
        if not f.filename: continue
        data = f.read()
        cur = db.execute(
            "INSERT INTO attachments(vessel_id,ref_type,ref_id,filename,filesize,mimetype,data)"
            " VALUES(?,?,?,?,?,?,?)",
            (vid, ref_type, ref_id, f.filename, len(data), f.content_type, data))
        uploaded.append({"id": cur.lastrowid, "filename": f.filename,
                         "filesize": len(data), "mimetype": f.content_type})
    db.commit()
    return jsonify(uploaded), 201

@app.route("/api/attachments/<int:aid>", methods=["GET"])
def download_attachment(aid):
    r = get_db().execute(
        "SELECT filename,mimetype,data FROM attachments WHERE id=?", (aid,)).fetchone()
    if not r: abort(404)
    return send_file(io.BytesIO(r["data"]),
                     mimetype=r["mimetype"] or "application/octet-stream",
                     as_attachment=True, download_name=r["filename"])

@app.route("/api/attachments/<int:aid>/preview", methods=["GET"])
def preview_attachment(aid):
    r = get_db().execute(
        "SELECT filename,mimetype,data FROM attachments WHERE id=?", (aid,)).fetchone()
    if not r: abort(404)
    return send_file(io.BytesIO(r["data"]),
                     mimetype=r["mimetype"] or "application/octet-stream",
                     as_attachment=False, download_name=r["filename"])

@app.route("/api/attachments/<int:aid>", methods=["DELETE"])
def delete_attachment(aid):
    db = get_db()
    db.execute("DELETE FROM attachments WHERE id=?", (aid,))
    db.commit()
    return jsonify({"deleted": aid})



# ══════════════════════════════════════════════════════════════
# DAILY TRACKING LOGS — 공통 CRUD 헬퍼
# ══════════════════════════════════════════════════════════════

def _tracking_get(table, vid, order="id"):
    DATE_COLS = {'start_date','end_date','open_date','close_date','completion_date',
                 'stop_date','date','bottom_plug_open','bottom_plug_close',
                 'last_updated','updated_at','uploaded_at'}
    rows = []
    for r in get_db().execute(f"SELECT * FROM {table} WHERE vessel_id=? ORDER BY {order}", (vid,)).fetchall():
        d = dict(r)
        for k, v in d.items():
            if k in DATE_COLS and v and isinstance(v, str) and ' ' in v:
                d[k] = v.split(' ')[0]  # 시간 제거
        rows.append(d)
    return jsonify(rows)

def _tracking_bulk(table, vid, items, insert_fn):
    db = get_db()
    db.execute(f"DELETE FROM {table} WHERE vessel_id=?", (vid,))
    for item in items:
        insert_fn(db, vid, item)
    db.commit()
    return jsonify([dict(r) for r in
        get_db().execute(f"SELECT * FROM {table} WHERE vessel_id=? ORDER BY id", (vid,)).fetchall()])

def _tracking_delete(table, rid):
    db = get_db()
    db.execute(f"DELETE FROM {table} WHERE id=?", (rid,))
    db.commit()
    return jsonify({"deleted": rid})

def _tracking_update(table, rid, update_sql, params):
    db = get_db()
    db.execute(update_sql, (*params, rid))
    db.commit()
    return jsonify(dict(get_db().execute(f"SELECT * FROM {table} WHERE id=?", (rid,)).fetchone()))


# ── Steel Repair ──────────────────────────────────────────────
@app.route("/api/vessels/<vid>/steel_repair", methods=["GET"])
def get_steel_repair(vid): return _tracking_get("steel_repair", vid)

@app.route("/api/vessels/<vid>/steel_repair", methods=["POST"])
def create_steel_repair(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("""INSERT INTO steel_repair
        (vessel_id,no,position_tank,frame_no,location_detail,type,
         length_l,width_w,thickness_t,steel_grade,new_weight,
         space_type,shape,priority,status,start_date,completion_date,
         remark,description)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (vid,
         d.get("no",""),
         d.get("position_tank",""), d.get("frame_no",""),
         d.get("location_detail",""), d.get("type",""),
         d.get("length_l",""), d.get("width_w",""), d.get("thickness_t",""),
         d.get("steel_grade",""), d.get("new_weight",""),
         d.get("space_type","Open Space"), d.get("shape","Flat"),
         d.get("priority","Normal"), d.get("status","Not Started"),
         d.get("start_date") or None, d.get("completion_date") or None,
         d.get("remark",""),
         d.get("description","")))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM steel_repair WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/steel_repair/<int:rid>", methods=["PUT"])
def update_steel_repair(rid):
    d = request.get_json(force=True)
    return _tracking_update("steel_repair", rid,
        """UPDATE steel_repair SET no=?,position_tank=?,frame_no=?,location_detail=?,
           type=?,length_l=?,width_w=?,thickness_t=?,steel_grade=?,new_weight=?,
           space_type=?,shape=?,remark=?,priority=?,status=?,
           start_date=?,completion_date=?,
           last_updated=datetime('now') WHERE id=?""",
        (d.get("no",""), d.get("position_tank",""), d.get("frame_no",""),
         d.get("location_detail",""), d.get("type",""),
         d.get("length_l",""), d.get("width_w",""), d.get("thickness_t",""),
         d.get("steel_grade",""), d.get("new_weight",""),
         d.get("space_type","Open Space"), d.get("shape","Flat"),
         d.get("remark",""),
         d.get("priority","Normal"), d.get("status","Not Started"),
         d.get("start_date") or None, d.get("completion_date") or None))

@app.route("/api/steel_repair/<int:rid>", methods=["DELETE"])
def delete_steel_repair(rid): return _tracking_delete("steel_repair", rid)

@app.route("/api/vessels/<vid>/steel_repair/bulk", methods=["PUT"])
def bulk_steel_repair(vid):
    def ins(db, vid, item):
        db.execute("""INSERT INTO steel_repair
            (vessel_id,no,position_tank,frame_no,location_detail,type,
             length_l,width_w,thickness_t,steel_grade,new_weight,
             space_type,shape,priority,status,start_date,completion_date,
             test_required,staging_m3,est_cost,actual_charged,remark)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (vid, item.get("no",""), item.get("position_tank",""),
             item.get("frame_no",""), item.get("location_detail",""), item.get("type",""),
             item.get("length_l",""), item.get("width_w",""), item.get("thickness_t",""),
             item.get("steel_grade",""), item.get("new_weight",""),
             item.get("space_type","Open Space"), item.get("shape","Flat"),
             item.get("priority","Normal"), item.get("status","Not Started"),
             item.get("start_date") or None, item.get("completion_date") or None,
             item.get("test_required","None"), item.get("staging_m3",""),
             item.get("est_cost",""), item.get("actual_charged",""), item.get("remark","")))
    return _tracking_bulk("steel_repair", vid, request.get_json(force=True), ins)


# ── Pipe Repair ───────────────────────────────────────────────
def _pipe_row(d):
    return (d.get("no",""), d.get("system_line",""), d.get("position_tank",""),
            d.get("frame_no",""), d.get("location_detail",""),
            d.get("pipe_od",""), d.get("schedule","Sch40"),
            d.get("material","Carbon Steel"),
            d.get("length_m",""), d.get("bend_qty",""), d.get("flange_qty",""),
            d.get("valve_type","None"), d.get("valve_size",""), d.get("valve_qty",""),
            d.get("space_type","Open Deck / Open Space"),
            d.get("removal_refit","Not Required"),
            d.get("priority","Normal"), d.get("status","Not Started"),
            d.get("start_date") or None, d.get("completion_date") or None,
            d.get("est_cost",""), d.get("actual_charged",""),
            d.get("description",""), d.get("remark",""))

@app.route("/api/vessels/<vid>/pipe_repair", methods=["GET"])
def get_pipe_repair(vid): return _tracking_get("pipe_repair", vid)

@app.route("/api/vessels/<vid>/pipe_repair", methods=["POST"])
def create_pipe_repair(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("""INSERT INTO pipe_repair
        (vessel_id,no,system_line,position_tank,frame_no,location_detail,
         pipe_od,schedule,material,length_m,bend_qty,flange_qty,
         valve_type,valve_size,valve_qty,space_type,removal_refit,
         priority,status,start_date,completion_date,est_cost,actual_charged,description,remark)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (vid,) + _pipe_row(d))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM pipe_repair WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/pipe_repair/<int:rid>", methods=["PUT"])
def update_pipe_repair(rid):
    d = request.get_json(force=True)
    return _tracking_update("pipe_repair", rid,
        """UPDATE pipe_repair SET no=?,system_line=?,position_tank=?,frame_no=?,location_detail=?,
           pipe_od=?,schedule=?,material=?,length_m=?,bend_qty=?,flange_qty=?,
           valve_type=?,valve_size=?,valve_qty=?,space_type=?,removal_refit=?,
           priority=?,status=?,start_date=?,completion_date=?,
           est_cost=?,actual_charged=?,description=?,remark=?,last_updated=datetime('now') WHERE id=?""",
        _pipe_row(d))

@app.route("/api/pipe_repair/<int:rid>", methods=["DELETE"])
def delete_pipe_repair(rid): return _tracking_delete("pipe_repair", rid)

@app.route("/api/vessels/<vid>/pipe_repair/bulk", methods=["PUT"])
def bulk_pipe_repair(vid):
    def ins(db, vid, item):
        db.execute("""INSERT INTO pipe_repair
            (vessel_id,no,system_line,position_tank,frame_no,location_detail,
             pipe_od,schedule,material,length_m,bend_qty,flange_qty,
             valve_type,valve_size,valve_qty,space_type,removal_refit,
             priority,status,start_date,completion_date,est_cost,actual_charged,description,remark)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (vid,) + _pipe_row(item))
    return _tracking_bulk("pipe_repair", vid, request.get_json(force=True), ins)


# ── Outfitting ────────────────────────────────────────────────
@app.route("/api/vessels/<vid>/outfitting", methods=["GET"])
def get_outfitting(vid): return _tracking_get("outfitting", vid)

@app.route("/api/vessels/<vid>/outfitting", methods=["POST"])
def create_outfitting(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("INSERT INTO outfitting(vessel_id,no,description,location,priority,status,start_date,completion_date,remark) VALUES(?,?,?,?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("description",""), d.get("location",""),
         d.get("priority","Normal"), d.get("status","Open"),
         d.get("start_date") or None, d.get("completion_date") or None, d.get("remark","")))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM outfitting WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/outfitting/<int:rid>", methods=["PUT"])
def update_outfitting(rid):
    d = request.get_json(force=True)
    return _tracking_update("outfitting", rid,
        "UPDATE outfitting SET no=?,description=?,location=?,priority=?,status=?,start_date=?,completion_date=?,remark=?,last_updated=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("description",""), d.get("location",""),
         d.get("priority","Normal"), d.get("status","Open"),
         d.get("start_date") or None, d.get("completion_date") or None, d.get("remark","")))

@app.route("/api/outfitting/<int:rid>", methods=["DELETE"])
def delete_outfitting(rid): return _tracking_delete("outfitting", rid)

@app.route("/api/vessels/<vid>/outfitting/bulk", methods=["PUT"])
def bulk_outfitting(vid):
    def ins(db, vid, item):
        db.execute("INSERT INTO outfitting(vessel_id,no,description,location,priority,status,start_date,completion_date,remark) VALUES(?,?,?,?,?,?,?,?,?)",
            (vid, item.get("no",""), item.get("description",""), item.get("location",""),
             item.get("priority","Normal"), item.get("status","Open"),
             item.get("start_date") or None, item.get("completion_date") or None, item.get("remark","")))
    return _tracking_bulk("outfitting", vid, request.get_json(force=True), ins)


# ── WBT & COT ─────────────────────────────────────────────────
@app.route("/api/vessels/<vid>/wbt_cot", methods=["GET"])
def get_wbt_cot(vid): return _tracking_get("wbt_cot", vid)

@app.route("/api/vessels/<vid>/wbt_cot", methods=["POST"])
def create_wbt_cot(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("INSERT INTO wbt_cot(vessel_id,no,tank_name,manhole_status,open_date,close_date,bottom_plug_open,bottom_plug_close,remark) VALUES(?,?,?,?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("tank_name",""), d.get("manhole_status",""),
         d.get("open_date") or None, d.get("close_date") or None,
         d.get("bottom_plug_open",""), d.get("bottom_plug_close",""), d.get("remark","")))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM wbt_cot WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/wbt_cot/<int:rid>", methods=["PUT"])
def update_wbt_cot(rid):
    d = request.get_json(force=True)
    return _tracking_update("wbt_cot", rid,
        "UPDATE wbt_cot SET no=?,tank_name=?,manhole_status=?,open_date=?,close_date=?,bottom_plug_open=?,bottom_plug_close=?,remark=?,updated_at=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("tank_name",""), d.get("manhole_status",""),
         d.get("open_date") or None, d.get("close_date") or None,
         d.get("bottom_plug_open",""), d.get("bottom_plug_close",""), d.get("remark","")))

@app.route("/api/wbt_cot/<int:rid>", methods=["DELETE"])
def delete_wbt_cot(rid): return _tracking_delete("wbt_cot", rid)

@app.route("/api/vessels/<vid>/wbt_cot/bulk", methods=["PUT"])
def bulk_wbt_cot(vid):
    def ins(db, vid, item):
        db.execute("INSERT INTO wbt_cot(vessel_id,no,tank_name,manhole_status,open_date,close_date,bottom_plug_open,bottom_plug_close,remark) VALUES(?,?,?,?,?,?,?,?,?)",
            (vid, item.get("no",""), item.get("tank_name",""), item.get("manhole_status",""),
             item.get("open_date") or None, item.get("close_date") or None,
             item.get("bottom_plug_open",""), item.get("bottom_plug_close",""), item.get("remark","")))
    return _tracking_bulk("wbt_cot", vid, request.get_json(force=True), ins)


# ── Portable Fan ──────────────────────────────────────────────
@app.route("/api/vessels/<vid>/portable_fan", methods=["GET"])
def get_portable_fan(vid): return _tracking_get("portable_fan", vid)

@app.route("/api/vessels/<vid>/portable_fan", methods=["POST"])
def create_portable_fan(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("INSERT INTO portable_fan(vessel_id,no,location,qty,start_date,stop_date,remark) VALUES(?,?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("location",""), d.get("qty",""),
         d.get("start_date") or None, d.get("stop_date") or None, d.get("remark","")))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM portable_fan WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/portable_fan/<int:rid>", methods=["PUT"])
def update_portable_fan(rid):
    d = request.get_json(force=True)
    return _tracking_update("portable_fan", rid,
        "UPDATE portable_fan SET no=?,location=?,qty=?,start_date=?,stop_date=?,remark=?,updated_at=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("location",""), d.get("qty",""),
         d.get("start_date") or None, d.get("stop_date") or None, d.get("remark","")))

@app.route("/api/portable_fan/<int:rid>", methods=["DELETE"])
def delete_portable_fan(rid): return _tracking_delete("portable_fan", rid)

@app.route("/api/vessels/<vid>/portable_fan/bulk", methods=["PUT"])
def bulk_portable_fan(vid):
    def ins(db, vid, item):
        db.execute("INSERT INTO portable_fan(vessel_id,no,location,qty,start_date,stop_date,remark) VALUES(?,?,?,?,?,?,?)",
            (vid, item.get("no",""), item.get("location",""), item.get("qty",""),
             item.get("start_date") or None, item.get("stop_date") or None, item.get("remark","")))
    return _tracking_bulk("portable_fan", vid, request.get_json(force=True), ins)


# ── Staging ───────────────────────────────────────────────────
@app.route("/api/vessels/<vid>/staging", methods=["GET"])
def get_staging(vid): return _tracking_get("staging", vid)

@app.route("/api/vessels/<vid>/staging", methods=["POST"])
def create_staging(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("INSERT INTO staging(vessel_id,no,location,staging_area,qty,remark) VALUES(?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("location",""), d.get("staging_area",""), d.get("qty",""), d.get("remark","")))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM staging WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/staging/<int:rid>", methods=["PUT"])
def update_staging(rid):
    d = request.get_json(force=True)
    return _tracking_update("staging", rid,
        "UPDATE staging SET no=?,location=?,staging_area=?,qty=?,remark=?,updated_at=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("location",""), d.get("staging_area",""), d.get("qty",""), d.get("remark","")))

@app.route("/api/staging/<int:rid>", methods=["DELETE"])
def delete_staging(rid): return _tracking_delete("staging", rid)

@app.route("/api/vessels/<vid>/staging/bulk", methods=["PUT"])
def bulk_staging(vid):
    def ins(db, vid, item):
        db.execute("INSERT INTO staging(vessel_id,no,location,staging_area,qty,remark) VALUES(?,?,?,?,?,?)",
            (vid, item.get("no",""), item.get("location",""), item.get("staging_area",""), item.get("qty",""), item.get("remark","")))
    return _tracking_bulk("staging", vid, request.get_json(force=True), ins)


# ── Gas Free ──────────────────────────────────────────────────
@app.route("/api/vessels/<vid>/gas_free", methods=["GET"])
def get_gas_free(vid): return _tracking_get("gas_free", vid)

@app.route("/api/vessels/<vid>/gas_free", methods=["POST"])
def create_gas_free(vid):
    d = request.get_json(force=True); db = get_db()
    cur = db.execute("INSERT INTO gas_free(vessel_id,no,tank,certificate,date,remark) VALUES(?,?,?,?,?,?)",
        (vid, d.get("no",""), d.get("tank",""), d.get("certificate",""), d.get("date") or None, d.get("remark","")))
    db.commit()
    return jsonify(dict(get_db().execute("SELECT * FROM gas_free WHERE id=?", (cur.lastrowid,)).fetchone())), 201

@app.route("/api/gas_free/<int:rid>", methods=["PUT"])
def update_gas_free(rid):
    d = request.get_json(force=True)
    return _tracking_update("gas_free", rid,
        "UPDATE gas_free SET no=?,tank=?,certificate=?,date=?,remark=?,updated_at=datetime('now') WHERE id=?",
        (d.get("no",""), d.get("tank",""), d.get("certificate",""), d.get("date") or None, d.get("remark","")))

@app.route("/api/gas_free/<int:rid>", methods=["DELETE"])
def delete_gas_free(rid): return _tracking_delete("gas_free", rid)

@app.route("/api/vessels/<vid>/gas_free/bulk", methods=["PUT"])
def bulk_gas_free(vid):
    def ins(db, vid, item):
        db.execute("INSERT INTO gas_free(vessel_id,no,tank,certificate,date,remark) VALUES(?,?,?,?,?,?)",
            (vid, item.get("no",""), item.get("tank",""), item.get("certificate",""), item.get("date") or None, item.get("remark","")))
    return _tracking_bulk("gas_free", vid, request.get_json(force=True), ins)


# ── Tank Plan ─────────────────────────────────────────────────
@app.route("/api/vessels/<vid>/tank_plan", methods=["GET"])
@login_required
def get_tank_plan(vid):
    items = [dict(r) for r in get_db().execute(
        """SELECT id, no, position_tank, frame_no, location_detail,
                  type, priority, status, description,
                  length_l, width_w, thickness_t, new_weight, space_type
           FROM steel_repair WHERE vessel_id=? ORDER BY id""",
        (vid,)).fetchall()]
    return jsonify(items)


@app.route("/api/vessels/<vid>/pipe_plan", methods=["GET"])
@login_required
def get_pipe_plan(vid):
    items = [dict(r) for r in get_db().execute(
        """SELECT id, no, system_line, position_tank, frame_no, location_detail,
                  pipe_od, schedule, material, length_m, bend_qty, flange_qty,
                  valve_type, valve_size, valve_qty, priority, status, description
           FROM pipe_repair WHERE vessel_id=? ORDER BY id""",
        (vid,)).fetchall()]
    return jsonify(items)


@app.route("/api/vessels/<vid>/tank_layout", methods=["GET"])
@login_required
def get_tank_layout(vid):
    r = get_db().execute(
        "SELECT layout_json FROM vessel_tank_layout WHERE vessel_id=?", (vid,)).fetchone()
    return jsonify(json.loads(r['layout_json'])) if r else jsonify(None)


@app.route("/api/vessels/<vid>/tank_layout", methods=["PUT"])
@login_required
@viewer_forbidden
def save_tank_layout(vid):
    layout_str = json.dumps(request.get_json(force=True))
    db = get_db()
    db.execute("""INSERT OR REPLACE INTO vessel_tank_layout(vessel_id, layout_json, updated_at)
                  VALUES(?, ?, datetime('now'))""", (vid, layout_str))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/vessels/<path:vid>/wps_criteria", methods=["GET"])
@login_required
def get_wps_criteria(vid):
    db = get_db()
    try:
        db.execute("""CREATE TABLE IF NOT EXISTS vessel_wps_criteria (
            vessel_id TEXT PRIMARY KEY, criteria_json TEXT NOT NULL,
            updated_at TEXT DEFAULT (datetime('now')))""")
        db.commit()
    except: pass
    row = db.execute(
        "SELECT criteria_json FROM vessel_wps_criteria WHERE vessel_id=?", (vid,)).fetchone()
    if not row: return jsonify(None)
    return jsonify(json.loads(row["criteria_json"]))

@app.route("/api/vessels/<path:vid>/wps_criteria", methods=["PUT"])
@login_required
@viewer_forbidden
def save_wps_criteria(vid):
    data = json.dumps(request.get_json(force=True))
    db = get_db()
    try:
        db.execute("""CREATE TABLE IF NOT EXISTS vessel_wps_criteria (
            vessel_id TEXT PRIMARY KEY, criteria_json TEXT NOT NULL,
            updated_at TEXT DEFAULT (datetime('now')))""")
    except: pass
    db.execute("""INSERT OR REPLACE INTO vessel_wps_criteria(vessel_id, criteria_json, updated_at)
                  VALUES(?, ?, datetime('now'))""", (vid, data))
    db.commit()
    return jsonify({"success": True})

@app.route("/api/vessels/<vid>/orphan_positions", methods=["GET"])
@login_required
def get_orphan_positions(vid):
    """현재 레이아웃에 없는 position_tank 값을 가진 항목 조회"""
    layout_row = get_db().execute(
        "SELECT layout_json FROM vessel_tank_layout WHERE vessel_id=?", (vid,)).fetchone()
    layout_names = set()
    if layout_row:
        import json as _json
        layout = _json.loads(layout_row['layout_json'])
        for sec in layout.get('sections', []):
            for col in sec.get('columns', []):
                for f in ('p','c','s'):
                    t = col.get(f)
                    if t and not t.get('empty') and t.get('name'):
                        layout_names.add(t['name'].strip())

    result = {}
    for table in ('steel_repair', 'pipe_repair'):
        rows = get_db().execute(
            f"SELECT id, no, position_tank, description, priority, status FROM {table} "
            f"WHERE vessel_id=? AND position_tank IS NOT NULL AND position_tank != ''",
            (vid,)).fetchall()
        for r in rows:
            pt = (r['position_tank'] or '').strip()
            if pt and pt not in layout_names:
                if pt not in result:
                    result[pt] = {'steel': [], 'pipe': []}
                key = 'steel' if table == 'steel_repair' else 'pipe'
                result[pt][key].append({
                    'id': r['id'], 'no': r['no'],
                    'description': r['description'] or '',
                    'priority': r['priority'], 'status': r['status']
                })
    return jsonify(result)


@app.route("/api/vessels/<vid>/position_rename", methods=["PUT"])
@login_required
@viewer_forbidden
def rename_position(vid):
    """레이아웃 탱크명 변경 시 steel_repair / pipe_repair의 position_tank 자동 업데이트"""
    renames = request.get_json(force=True)  # [{old: "DBT 1P", new: "WBT 1P"}, ...]
    if not isinstance(renames, list) or not renames:
        return jsonify({"success": True, "updated": 0})
    db = get_db()
    total = 0
    for r in renames:
        old_name = r.get("old", "").strip()
        new_name = r.get("new", "").strip()
        if not old_name or not new_name or old_name == new_name:
            continue
        for table in ("steel_repair", "pipe_repair"):
            cur = db.execute(
                f"UPDATE {table} SET position_tank=? WHERE vessel_id=? AND position_tank=?",
                (new_name, vid, old_name))
            total += cur.rowcount
    db.commit()
    return jsonify({"success": True, "updated": total})


@app.route("/api/tracking/template")
def download_tracking_template():
    """Daily Tracking Log 템플릿 xlsx 다운로드"""
    template_path = os.path.join(os.path.dirname(__file__), "static", "templates", "DD_DAILY_LOG_TEMPLATE_REV2.xlsx")
    if not os.path.exists(template_path):
        # fallback to old template name
        template_path = os.path.join(os.path.dirname(__file__), "static", "templates", "DD_DAILY_LOG_TEMPLATE.xlsx")
    return send_file(template_path,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="DD_DAILY_LOG_TEMPLATE_REV2.xlsx")


# ── XLSX 통합 업로드 ──────────────────────────────────────────
@app.route("/api/vessels/<vid>/tracking/upload_xlsx", methods=["POST"])
def upload_tracking_xlsx(vid):
    """xlsx 파일을 받아 각 시트별로 DB에 저장"""
    import openpyxl, io as _io

    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "xlsx 파일만 가능합니다"}), 400

    wb = openpyxl.load_workbook(_io.BytesIO(f.read()), read_only=True, data_only=True)
    db = get_db()
    result = {}

    def val(v): return str(v).strip() if v is not None else ""

    # Priority / Status 값 정규화
    PRI_MAP = {
        'normal':'Normal','urgent':'Urgent','critical':'Critical','on hold':'On Hold',
        'hold':'On Hold',
        'low':'Low','medium':'Medium','high':'High',  # steel/outfit용
    }
    STAT_MAP = {
        'open':'Not Started','not started':'Not Started',
        'in progress':'In Progress','inprogress':'In Progress',
        'completed':'Completed','done':'Completed','close':'Completed','closed':'Completed',
        'on hold':'On Hold','hold':'On Hold',
    }
    def norm_pri(v): return PRI_MAP.get(val(v).lower(), val(v) or 'Normal')
    def norm_stat(v): return STAT_MAP.get(val(v).lower(), val(v) or 'Not Started')

    def get_data_rows(ws, header_row=4):
        rows = list(ws.iter_rows(min_row=header_row+1, values_only=True))
        return [r for r in rows if any(v is not None and str(v).strip() for v in r)]

    # Steel Repair — xlsx 자동 업로드 제외 (직접 입력 전용)
    # Pipe Repair — xlsx 자동 업로드 제외 (직접 입력 전용)

    # ── Upsert 헬퍼 ──────────────────────────────────────────────
    # 시트가 비어있으면 → 기존 데이터 유지 (skip)
    # No. 일치 → UPDATE, 신규 No. → INSERT
    def upsert_tracking(sheet_name, table, no_col, data_rows_fn, insert_fn, update_fn):
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        rows_in_sheet = data_rows_fn(ws)
        valid_rows = [r for r in rows_in_sheet if any(v for v in r)]
        if not valid_rows:
            return 0  # 시트 비어있음 → 기존 데이터 유지, 0건 처리
        # 기존 No. 목록 조회
        existing = {r[no_col]: r['id']
                    for r in db.execute(
                        f"SELECT id, {no_col} FROM {table} WHERE vessel_id=? AND {no_col} IS NOT NULL AND {no_col}!=''",
                        (vid,)).fetchall()}
        cnt_upsert = 0
        for r in valid_rows:
            no_val = insert_fn(r, existing, db, vid)
            if no_val:
                cnt_upsert += 1
        return cnt_upsert

    # Outfitting Daily Log
    # Template: A(0)=No, B(1)=Description, C(2)=Location, D(3)=Priority,
    #           E(4)=Start Date, F(5)=Completion, G(6)=Status, H(7)=Remark
    if "Outfitting Daily Log" in wb.sheetnames:
        ws = wb["Outfitting Daily Log"]
        sheet_rows = get_data_rows(ws, 4)
        valid_rows = [r for r in sheet_rows if any(v for v in r)]
        if valid_rows:
            existing = {str(r['no']): r['id'] for r in db.execute(
                "SELECT id, no FROM outfitting WHERE vessel_id=? AND no IS NOT NULL AND no!=''", (vid,)).fetchall()}
            cnt = 0
            for r in valid_rows:
                no_val = val(r[0])
                fields = (val(r[1]), val(r[2]), norm_pri(val(r[3])),
                          norm_stat(val(r[6]) if len(r)>6 else ''),
                          val(r[4]) or None, val(r[5]) or None,
                          val(r[7]) if len(r)>7 else '')
                if no_val and no_val in existing:
                    db.execute("""UPDATE outfitting SET description=?,location=?,priority=?,status=?,
                        start_date=?,completion_date=?,remark=?,last_updated=datetime('now')
                        WHERE id=?""", fields + (existing[no_val],))
                else:
                    db.execute("INSERT INTO outfitting(vessel_id,no,description,location,priority,status,start_date,completion_date,remark) VALUES(?,?,?,?,?,?,?,?,?)",
                        (vid, no_val) + fields)
                cnt += 1
            result["outfitting"] = cnt
        else:
            result["outfitting"] = 0

    # WBT & COT
    # Template: A(0)=AutoNo, B(1)=No, C(2)=Tank, D(3)=Manhole,
    #           E(4)=OpenDate, F(5)=CloseDate, G(6)=BottomPlugOpen, H(7)=BottomPlugClose, I(8)=Remark
    if "WBT & COT" in wb.sheetnames:
        ws = wb["WBT & COT"]
        sheet_rows = get_data_rows(ws, 4)
        valid_rows = [r for r in sheet_rows if len(r)>=3 and any(v for v in r)]
        if valid_rows:
            existing = {str(r['no']): r['id'] for r in db.execute(
                "SELECT id, no FROM wbt_cot WHERE vessel_id=? AND no IS NOT NULL AND no!=''", (vid,)).fetchall()}
            cnt = 0
            for r in valid_rows:
                no_val = val(r[1])
                fields = (val(r[2]), val(r[3]),
                          val(r[4]) or None, val(r[5]) or None,
                          val(r[6]), val(r[7]),
                          val(r[8]) if len(r)>8 else '')
                if no_val and no_val in existing:
                    db.execute("""UPDATE wbt_cot SET tank_name=?,manhole_status=?,open_date=?,close_date=?,
                        bottom_plug_open=?,bottom_plug_close=?,remark=?,updated_at=datetime('now')
                        WHERE id=?""", fields + (existing[no_val],))
                else:
                    db.execute("INSERT INTO wbt_cot(vessel_id,no,tank_name,manhole_status,open_date,close_date,bottom_plug_open,bottom_plug_close,remark) VALUES(?,?,?,?,?,?,?,?,?)",
                        (vid, no_val) + fields)
                cnt += 1
            result["wbt_cot"] = cnt
        else:
            result["wbt_cot"] = 0

    # Portable Fan Installation
    # Template: A(0)=AutoNo, B(1)=No, C(2)=Location, D(3)=Qty, E(4)=StartDate, F(5)=StopDate, G(6)=Remark
    if "Portable Fan Installation" in wb.sheetnames:
        ws = wb["Portable Fan Installation"]
        sheet_rows = get_data_rows(ws, 4)
        valid_rows = [r for r in sheet_rows if len(r)>=3 and any(v for v in r)]
        if valid_rows:
            existing = {str(r['no']): r['id'] for r in db.execute(
                "SELECT id, no FROM portable_fan WHERE vessel_id=? AND no IS NOT NULL AND no!=''", (vid,)).fetchall()}
            cnt = 0
            for r in valid_rows:
                no_val = val(r[1])
                fields = (val(r[2]), val(r[3]),
                          val(r[4]) or None, val(r[5]) or None,
                          val(r[6]) if len(r)>6 else '')
                if no_val and no_val in existing:
                    db.execute("""UPDATE portable_fan SET location=?,qty=?,start_date=?,stop_date=?,
                        remark=?,updated_at=datetime('now') WHERE id=?""", fields + (existing[no_val],))
                else:
                    db.execute("INSERT INTO portable_fan(vessel_id,no,location,qty,start_date,stop_date,remark) VALUES(?,?,?,?,?,?,?)",
                        (vid, no_val) + fields)
                cnt += 1
            result["portable_fan"] = cnt
        else:
            result["portable_fan"] = 0

    # Staging
    # Template: A(0)=AutoNo, B(1)=No, C(2)=Location/Frame, D(3)=StagingArea, E(4)=Qty, F(5)=Remark
    if "Staging" in wb.sheetnames:
        ws = wb["Staging"]
        sheet_rows = get_data_rows(ws, 4)
        valid_rows = [r for r in sheet_rows if len(r)>=3 and any(v for v in r)]
        if valid_rows:
            existing = {str(r['no']): r['id'] for r in db.execute(
                "SELECT id, no FROM staging WHERE vessel_id=? AND no IS NOT NULL AND no!=''", (vid,)).fetchall()}
            cnt = 0
            for r in valid_rows:
                no_val = val(r[1])
                fields = (val(r[2]), val(r[3]), val(r[4]),
                          val(r[5]) if len(r)>5 else '')
                if no_val and no_val in existing:
                    db.execute("""UPDATE staging SET location=?,staging_area=?,qty=?,
                        remark=?,updated_at=datetime('now') WHERE id=?""", fields + (existing[no_val],))
                else:
                    db.execute("INSERT INTO staging(vessel_id,no,location,staging_area,qty,remark) VALUES(?,?,?,?,?,?)",
                        (vid, no_val) + fields)
                cnt += 1
            result["staging"] = cnt
        else:
            result["staging"] = 0

    # Gas Free Certificate
    # Template: A(0)=AutoNo, B(1)=No, C(2)=Tank, D(3)=Certificate, E(4)=Date, F(5)=Remark
    if "Gas Free Certificate" in wb.sheetnames:
        ws = wb["Gas Free Certificate"]
        sheet_rows = get_data_rows(ws, 4)
        valid_rows = [r for r in sheet_rows if len(r)>=3 and any(v for v in r)]
        if valid_rows:
            existing = {str(r['no']): r['id'] for r in db.execute(
                "SELECT id, no FROM gas_free WHERE vessel_id=? AND no IS NOT NULL AND no!=''", (vid,)).fetchall()}
            cnt = 0
            for r in valid_rows:
                no_val = val(r[1])
                fields = (val(r[2]), val(r[3]),
                          val(r[4]) or None,
                          val(r[5]) if len(r)>5 else '')
                if no_val and no_val in existing:
                    db.execute("""UPDATE gas_free SET tank=?,certificate=?,date=?,
                        remark=?,updated_at=datetime('now') WHERE id=?""", fields + (existing[no_val],))
                else:
                    db.execute("INSERT INTO gas_free(vessel_id,no,tank,certificate,date,remark) VALUES(?,?,?,?,?,?)",
                        (vid, no_val) + fields)
                cnt += 1
            result["gas_free"] = cnt
        else:
            result["gas_free"] = 0

    db.commit()
    return jsonify({"success": True, "imported": result}), 201


# ══ VESSEL DOCUMENTS ══════════════════════════════════════════

DOC_TYPES = ['Shipyard Specification', 'Shipyard Quotation', 'Shipyard Workdone List',
             'Yard Report', 'Service Report', 'Invoices', 'Reference']

@app.route("/api/vessels/<vid>/documents", methods=["GET"])
@login_required
def get_documents(vid):
    import re
    def natural_key(item):
        return [int(c) if c.isdigit() else c.lower()
                for c in re.split(r'(\d+)', item.get('filename',''))]

    doc_type = request.args.get('doc_type')
    if doc_type:
        data = rows("SELECT id,doc_type,filename,filesize,mimetype,uploaded_at FROM vessel_documents WHERE vessel_id=? AND doc_type=?", vid, doc_type)
    else:
        data = rows("SELECT id,doc_type,filename,filesize,mimetype,uploaded_at FROM vessel_documents WHERE vessel_id=?", vid)
    data.sort(key=lambda x: (x.get('doc_type',''), natural_key(x)))
    return jsonify(data)

@app.route("/api/vessels/<vid>/documents", methods=["POST"])
@login_required
@viewer_forbidden
def upload_document(vid):
    doc_type = request.form.get('doc_type','')
    if doc_type not in DOC_TYPES:
        return jsonify({"error": "Invalid doc_type"}), 400
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    db = get_db()
    uploaded = []
    for f in request.files.getlist("file"):
        if not f.filename: continue
        data = f.read()
        cur = db.execute(
            "INSERT INTO vessel_documents(vessel_id,doc_type,filename,filesize,mimetype,data) VALUES(?,?,?,?,?,?)",
            (vid, doc_type, f.filename, len(data), f.content_type, data))
        uploaded.append({"id": cur.lastrowid, "doc_type": doc_type, "filename": f.filename,
                         "filesize": len(data), "mimetype": f.content_type})
    db.commit()
    return jsonify(uploaded), 201

@app.route("/api/documents/<int:did>", methods=["GET"])
@login_required
def download_document(did):
    r = get_db().execute("SELECT filename,mimetype,data FROM vessel_documents WHERE id=?", (did,)).fetchone()
    if not r: abort(404)
    return send_file(io.BytesIO(r["data"]), mimetype=r["mimetype"] or "application/octet-stream",
                     as_attachment=True, download_name=r["filename"])

@app.route("/api/documents/<int:did>/preview", methods=["GET"])
@login_required
def preview_document(did):
    r = get_db().execute("SELECT filename,mimetype,data FROM vessel_documents WHERE id=?", (did,)).fetchone()
    if not r: abort(404)
    return send_file(io.BytesIO(r["data"]), mimetype=r["mimetype"] or "application/octet-stream",
                     as_attachment=False, download_name=r["filename"])


@app.route("/api/documents/<int:did>", methods=["DELETE"])
@login_required
@viewer_forbidden
def delete_document(did):
    db = get_db()
    db.execute("DELETE FROM vessel_documents WHERE id=?", (did,))
    db.commit()
    return jsonify({"deleted": did})


# ══════════════════════════════════════════════════════════════
# MCP CONNECTOR v3 (Claude 커넥터용 OAuth 2.0 + MCP 엔드포인트)
# ── Streamable-HTTP + RFC 7591 DCR 지원 ──────────────────────
#   ✔ Dynamic Client Registration (/oauth/register) 추가
#   ✔ registration_endpoint 메타데이터 광고
#   ✔ initialize / notifications/initialized / ping 처리
#   ✔ CORS preflight + 필수 헤더 노출
#   ✔ GET / DELETE 메서드 표준 처리 (405 / 204)
#   ✔ tools/call 예외 안전화 (isError 응답)
# ══════════════════════════════════════════════════════════════
import secrets, time, hashlib, base64
from flask import stream_with_context, Response

# ── CORS (MCP / OAuth / 메타데이터 엔드포인트만) ────────────────
try:
    from flask_cors import CORS
    CORS(app,
         resources={
             r"/mcp":           {"origins": "*"},
             r"/oauth/*":       {"origins": "*"},
             r"/.well-known/*": {"origins": "*"},
         },
         supports_credentials=False,
         expose_headers=["WWW-Authenticate", "Mcp-Session-Id"],
         allow_headers=["Authorization", "Content-Type",
                        "Mcp-Session-Id", "MCP-Protocol-Version",
                        "Accept"])
except ImportError:
    pass  # flask-cors 미설치여도 서버는 뜨도록

_mcp_tokens  = {}   # token     -> {"username":…, "exp":…, "client_id":…}
_auth_codes  = {}   # code      -> {"username":…, "exp":…, "code_challenge":…, …}
_oauth_clients = {} # client_id -> {redirect_uris, client_name, created_at, …}

# MCP 최신 stable protocol version
MCP_PROTOCOL_VERSION = "2025-06-18"
MCP_SERVER_NAME      = "DD Manager"
MCP_SERVER_VERSION   = "1.0.0"


def _authenticate(username, password):
    """기존 DB 인증 재사용 — hash_pw는 앱 init 단계에서 module-level로 정의되어 있음"""
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE username=? AND password_hash=?",
        (username, hash_pw(password))
    ).fetchone()
    return dict(user) if user else None


# ── OAuth 2.0 Authorization Server 메타데이터 (RFC 8414) ────
@app.route('/.well-known/oauth-authorization-server', methods=['GET'])
def oauth_metadata():
    base = request.headers.get('X-Forwarded-Proto','https') + '://' + request.host
    return jsonify({
        "issuer": base,
        "authorization_endpoint":           f"{base}/oauth/authorize",
        "token_endpoint":                   f"{base}/oauth/token",
        "registration_endpoint":            f"{base}/oauth/register",   # ★ DCR
        "response_types_supported":         ["code"],
        "grant_types_supported":            ["authorization_code", "refresh_token"],
        "code_challenge_methods_supported": ["S256"],
        "token_endpoint_auth_methods_supported": ["none", "client_secret_post"],
        "scopes_supported":                 ["mcp"]
    })


@app.route('/.well-known/oauth-protected-resource', methods=['GET'])
def oauth_protected_resource():
    base = request.headers.get('X-Forwarded-Proto','https') + '://' + request.host
    return jsonify({
        "resource": f"{base}/mcp",
        "authorization_servers": [base],
        "scopes_supported": ["mcp"],
        "bearer_methods_supported": ["header"]
    })


# ── OAuth 2.0 Dynamic Client Registration (RFC 7591) ────────
@app.route('/oauth/register', methods=['POST','OPTIONS'])
def oauth_register():
    if request.method == 'OPTIONS':
        return ('', 204)

    data = request.get_json(force=True, silent=True) or {}

    client_id     = secrets.token_urlsafe(24)
    client_secret = secrets.token_urlsafe(32)
    now = int(time.time())

    client_info = {
        "client_id":     client_id,
        "client_secret": client_secret,
        "client_name":   data.get("client_name", "Unnamed MCP Client"),
        "redirect_uris": data.get("redirect_uris", []),
        "grant_types":   data.get("grant_types",    ["authorization_code"]),
        "response_types":data.get("response_types", ["code"]),
        "token_endpoint_auth_method":
            data.get("token_endpoint_auth_method", "none"),
        "scope":         data.get("scope", "mcp"),
        "created_at":    now
    }
    _oauth_clients[client_id] = client_info

    resp = jsonify({
        "client_id":                client_id,
        "client_secret":            client_secret,
        "client_id_issued_at":      now,
        "client_secret_expires_at": 0,        # 0 = never expires
        "client_name":              client_info["client_name"],
        "redirect_uris":            client_info["redirect_uris"],
        "grant_types":              client_info["grant_types"],
        "response_types":           client_info["response_types"],
        "token_endpoint_auth_method": client_info["token_endpoint_auth_method"],
        "scope":                    client_info["scope"]
    })
    resp.status_code = 201
    return resp


# ── OAuth 인가 엔드포인트 ────────────────────────────────────
@app.route('/oauth/authorize', methods=['GET','POST'])
def oauth_authorize():
    if request.method == 'GET':
        client_id             = request.args.get('client_id','')
        redirect_uri          = request.args.get('redirect_uri','')
        state                 = request.args.get('state','')
        code_challenge        = request.args.get('code_challenge','')
        code_challenge_method = request.args.get('code_challenge_method','S256')
        scope                 = request.args.get('scope','mcp')
        return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>DD Manager 로그인</title>
<style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;height:100vh;margin:0;background:#0d1b2a}}
.box{{background:#1a2a3a;padding:2rem;border-radius:12px;width:340px;color:#fff}}
h2{{margin-top:0}}.client{{font-size:.85rem;opacity:.7;margin-bottom:1rem}}
input{{width:100%;padding:.6rem;margin:.4rem 0 1rem;box-sizing:border-box;border-radius:6px;border:1px solid #555;background:#0d1b2a;color:#fff}}
button{{width:100%;padding:.7rem;background:#3b82f6;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:1rem}}</style></head>
<body><div class="box"><h2>🚢 DD Manager</h2>
<div class="client">client: {client_id[:16] if client_id else '(none)'}…</div>
<form method="POST">
  <input type="hidden" name="client_id"              value="{client_id}">
  <input type="hidden" name="redirect_uri"           value="{redirect_uri}">
  <input type="hidden" name="state"                  value="{state}">
  <input type="hidden" name="code_challenge"         value="{code_challenge}">
  <input type="hidden" name="code_challenge_method"  value="{code_challenge_method}">
  <input type="hidden" name="scope"                  value="{scope}">
  <label>아이디</label><input name="username" required>
  <label>비밀번호</label><input type="password" name="password" required>
  <button type="submit">로그인 &amp; 연결</button>
</form></div></body></html>"""

    client_id             = request.form.get('client_id','')
    username              = request.form.get('username','').strip()
    password              = request.form.get('password','').strip()
    redirect_uri          = request.form.get('redirect_uri','')
    state                 = request.form.get('state','')
    code_challenge        = request.form.get('code_challenge','')
    code_challenge_method = request.form.get('code_challenge_method','S256')
    scope                 = request.form.get('scope','mcp')

    user = _authenticate(username, password)
    if not user:
        return "인증 실패. 뒤로 가서 다시 시도하세요.", 401

    code = secrets.token_urlsafe(32)
    _auth_codes[code] = {
        "client_id":             client_id,
        "username":              username,
        "redirect_uri":          redirect_uri,
        "scope":                 scope,
        "exp":                   time.time() + 300,
        "code_challenge":        code_challenge,
        "code_challenge_method": code_challenge_method,
    }
    sep = '&' if '?' in redirect_uri else '?'
    return redirect(f"{redirect_uri}{sep}code={code}&state={state}")


@app.route('/oauth/token', methods=['POST','OPTIONS'])
def oauth_token():
    if request.method == 'OPTIONS':
        return ('', 204)

    if request.is_json:
        data = request.get_json(force=True)
    else:
        data = request.form

    grant_type    = data.get('grant_type','authorization_code')

    if grant_type == 'authorization_code':
        code          = data.get('code','')
        code_verifier = data.get('code_verifier','')
        client_id     = data.get('client_id','')
        entry         = _auth_codes.pop(code, None)
        if not entry or time.time() > entry['exp']:
            return jsonify({"error":"invalid_grant"}), 400

        # PKCE 검증
        challenge = entry.get('code_challenge','')
        method    = entry.get('code_challenge_method','S256')
        if challenge:
            if method == 'S256':
                digest = base64.urlsafe_b64encode(
                    hashlib.sha256(code_verifier.encode()).digest()
                ).rstrip(b'=').decode()
                if digest != challenge:
                    return jsonify({
                        "error":"invalid_grant",
                        "error_description":"PKCE verification failed"
                    }), 400

        access_token  = secrets.token_urlsafe(48)
        refresh_token = secrets.token_urlsafe(48)
        _mcp_tokens[access_token] = {
            "username":  entry['username'],
            "client_id": entry.get('client_id') or client_id,
            "exp":       time.time() + 86400*30,
            "refresh":   refresh_token
        }
        return jsonify({
            "access_token":  access_token,
            "token_type":    "bearer",
            "expires_in":    86400*30,
            "refresh_token": refresh_token,
            "scope":         entry.get('scope','mcp')
        })

    if grant_type == 'refresh_token':
        refresh = data.get('refresh_token','')
        # 기존 토큰 중 refresh_token 일치하는 것 찾기
        for tok, info in list(_mcp_tokens.items()):
            if info.get('refresh') == refresh:
                _mcp_tokens.pop(tok, None)
                new_tok = secrets.token_urlsafe(48)
                new_ref = secrets.token_urlsafe(48)
                _mcp_tokens[new_tok] = {
                    "username":  info['username'],
                    "client_id": info.get('client_id',''),
                    "exp":       time.time() + 86400*30,
                    "refresh":   new_ref
                }
                return jsonify({
                    "access_token":  new_tok,
                    "token_type":    "bearer",
                    "expires_in":    86400*30,
                    "refresh_token": new_ref
                })
        return jsonify({"error":"invalid_grant"}), 400

    return jsonify({"error":"unsupported_grant_type"}), 400


# ── MCP 엔드포인트 ──────────────────────────────────────────
def _check_mcp_auth():
    auth  = request.headers.get('Authorization','')
    token = auth.replace('Bearer ','').strip()
    entry = _mcp_tokens.get(token)
    if not entry or time.time() > entry['exp']:
        return None
    return entry['username']


def _rpc_ok(req_id, result):
    return jsonify({"jsonrpc":"2.0","id":req_id,"result":result})


def _rpc_err(req_id, code, message, status=400):
    resp = jsonify({"jsonrpc":"2.0","id":req_id,
                    "error":{"code":code,"message":message}})
    resp.status_code = status
    return resp


@app.route('/mcp', methods=['GET','POST','DELETE','OPTIONS'])
def mcp_endpoint():
    # CORS preflight
    if request.method == 'OPTIONS':
        return ('', 204)

    # GET: 현재 SSE 미지원 → 405
    if request.method == 'GET':
        return jsonify({"error":"Method Not Allowed",
                        "hint":"Use POST with JSON-RPC 2.0 body"}), 405

    # DELETE: 세션 종료(옵션) → 그냥 204
    if request.method == 'DELETE':
        return ('', 204)

    # ── POST: 인증 필요 ───────────────────────────────────────
    username = _check_mcp_auth()
    if not username:
        base = request.headers.get('X-Forwarded-Proto','https') + '://' + request.host
        resp = jsonify({"error":"unauthorized"})
        resp.status_code = 401
        resp.headers['WWW-Authenticate'] = (
            f'Bearer realm="{base}",'
            f' resource_metadata_url="{base}/.well-known/oauth-protected-resource"'
        )
        return resp

    body   = request.get_json(force=True, silent=True) or {}
    method = body.get('method','')
    req_id = body.get('id')

    # 1) initialize 핸드셰이크
    if method == 'initialize':
        params = body.get('params',{})
        client_proto = params.get('protocolVersion', MCP_PROTOCOL_VERSION)
        return _rpc_ok(req_id, {
            "protocolVersion": client_proto,
            "capabilities": {
                "tools": {"listChanged": False}
            },
            "serverInfo": {
                "name":    MCP_SERVER_NAME,
                "version": MCP_SERVER_VERSION
            }
        })

    # 2) notifications/initialized (알림 → 204)
    if method == 'notifications/initialized':
        return ('', 204)

    # 3) ping (keep-alive)
    if method == 'ping':
        return _rpc_ok(req_id, {})

    # 4) tools/list
    if method == 'tools/list':
        tools = [
            {"name":"get_fleet_summary",
             "description":"전체 선단 현황 요약",
             "inputSchema":{"type":"object","properties":{}}},
            {"name":"get_vessels",
             "description":"등록된 모든 선박 목록 조회",
             "inputSchema":{"type":"object","properties":{}}},
            {"name":"get_vessel_jobs",
             "description":"특정 선박의 작업 목록 조회",
             "inputSchema":{
                 "type":"object",
                 "properties":{"vessel_id":{"type":"string","description":"선박 ID"}},
                 "required":["vessel_id"]
             }},
            {"name":"get_class_items",
             "description":"선급 검사 항목 조회",
             "inputSchema":{
                 "type":"object",
                 "properties":{"vessel_id":{"type":"string"}},
                 "required":["vessel_id"]
             }},
        ]
        return _rpc_ok(req_id, {"tools": tools})

    # 5) tools/call
    if method == 'tools/call':
        tool = body.get('params',{}).get('name')
        args = body.get('params',{}).get('arguments',{})
        db   = get_db()
        try:
            if tool == 'get_vessels':
                content = [dict(r) for r in db.execute(
                    "SELECT id,name,type,imo,shipyard,dock_in,dock_out FROM vessels"
                ).fetchall()]
            elif tool == 'get_vessel_jobs':
                vid = args.get('vessel_id','')
                content = [dict(r) for r in db.execute(
                    "SELECT number,section,category,description,status,completion "
                    "FROM jobs WHERE vessel_id=? ORDER BY number", (vid,)
                ).fetchall()]
            elif tool == 'get_class_items':
                vid = args.get('vessel_id','')
                content = [dict(r) for r in db.execute(
                    "SELECT * FROM class_items WHERE vessel_id=?", (vid,)
                ).fetchall()]
            elif tool == 'get_fleet_summary':
                vessels = [dict(r) for r in db.execute(
                    "SELECT id,name,type,imo,shipyard,dock_in,dock_out FROM vessels"
                ).fetchall()]
                content = {"vessel_count": len(vessels), "vessels": vessels}
            else:
                return _rpc_err(req_id, -32601, f"Unknown tool: {tool}", 404)
        except Exception as e:
            return _rpc_ok(req_id, {
                "content":[{"type":"text","text":f"Error: {e}"}],
                "isError": True
            })

        return _rpc_ok(req_id, {
            "content":[{"type":"text",
                        "text":json.dumps(content, ensure_ascii=False, default=str)}]
        })

    # 6) 알 수 없는 메서드
    if req_id is None:
        return ('', 204)
    return _rpc_err(req_id, -32601, f"Unknown method: {method}", 400)


# ── Run ───────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5000)
